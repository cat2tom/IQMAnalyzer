import pandas as pd

import os

from openpyxl import load_workbook


from collections import OrderedDict


def get_tps_patient_info(psqa_excel_file,work_sheet='Worksheet',tps_cell='L4',mrn_cell='D6',family_name_cell='D4',given_name_cell='D5'):
    
    """
    Read TPS and patient info from PSQA excel sheet.
    
    Input:
    
    output: 
    
    """
    
    pat_info=OrderedDict()
     
    wook_book=load_workbook(psqa_excel_file)
    
    if 'Worksheet' in wook_book.sheetnames:
    
        sheet=wook_book['Worksheet']
        
        tps=sheet[tps_cell].value
        
        mrn=sheet[mrn_cell].value
        
        family_name=sheet[family_name_cell].value
        
        given_name=sheet[given_name_cell].value
        
        pat_info['family_name']=family_name
        
        pat_info['given_name']=given_name
        
        pat_info['mrn']=mrn
        
        pat_info['tps']=tps
        
    else:
        
        pat_info['family_name']=''
               
        pat_info['given_name']=''
        
        pat_info['mrn']=''
        
        pat_info['tps']=''     
 
    
    
    return pat_info

def get_psqa_excel_reports(psqa_result_folder): 
    
    """
    Get all PSQA excel report under PSQA_result_folder recursivley 
    
    """
    
    all_psqa_reports=[]
    
    for dirpath, dirs, files in os.walk(psqa_result_folder): 
        for filename in files:
          
            fname = os.path.join(dirpath,filename)
                 
            if fname.endswith('.xlsm') and '_' in fname and ('~$' not in fname) :
                
                print('Processing '+fname)
                
                pat_inf=get_tps_patient_info(fname)
                
               
                   
                all_psqa_reports.extend([fname])
          
    return all_psqa_reports

def process_psqa_excel_reports(psqa_result_folder,process_fun):
    
    """
    Process PSQA excel reports using given function which take a excel file name as only input. 
    the outputs were 
    
    
    """
    
    collected_result=[]
    
      
    for dirpath, dirs, files in os.walk(psqa_result_folder): 
        for filename in files:
          
            fname = os.path.join(dirpath,filename)
                 
            if fname.endswith('.xlsm') and '_' in fname and ('~$' not in fname) :
                
                print('Processing '+fname)
                
                             
                pat_inf=process_fun(fname)
                
                                
                collected_result.extend([pat_inf])
                
                              
                            
    return collected_result
   
    
    
    
def read_patinf2csv(psqa_result_folder,process_fun=get_tps_patient_info,out_file='tps_mrn_list.csv'):
    
    """
    Read patient info into csv file.
    """
    
    patinfo_dic_list= process_psqa_excel_reports(psqa_result_folder,process_fun)
    
    pat_info_df=pd.DataFrame.from_dict(patinfo_dic_list)
    
    pat_info_df.to_csv(out_file)
    
    print('Processed results were written to '+out_file)



if __name__=='__main__':
    
    
    
    psqa_excel_file=r'C:\AitangResearch\IQMAnalyzer\testData\psqaSheet\KING_1512517_PH1_LTGROIN_PSQA.xlsm'
    
    #psqa_result_folder=r'V:\CTC-LiverpoolOncology-Physics\IMRT\PatientQA\Results'
    
    
    
    # # 1
    
    #pat_inf=get_tps_patient_info(psqa_excel_file)
    
    #print(pat_inf)
    
    # # 2
        
    #all_psqa_report=get_psqa_excel_reports(psqa_result_folder)
    
    #print(all_psqa_report)
    
    # # 3
    
    #process_fun=get_tps_patient_info
    
    #process_psqa_excel_reports(psqa_result_folder,process_fun)
    
    
    # # 3
    
    psqa_result_folder=r'V:\CTC-LiverpoolOncology-Physics\IMRT\PatientQA\2021\IMRT'
    
    read_patinf2csv(psqa_result_folder) 
    
    
    
    
    


    
    
    
    
    
    






